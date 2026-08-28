"""AI 보고서 양식 초안 — 추출 항목·샘플값·제목 위계를 보고 AI가 양식 설계를 제안하면,
openpyxl 로 {자리표시자} 초안 xlsx 를 조립한다.

흐름: 초안 생성 → 사용자가 다운로드해 엑셀에서 편집 → 다시 업로드 → 데이터 채워 최종본.
AI 는 '설계(JSON)'만 담당하고 파일 조립은 이 코드가 결정적으로 수행한다
(AI 가 엑셀을 직접 만들지 않으므로 깨진 파일이 나올 수 없고, 단위 테스트 가능).
"""
from __future__ import annotations

import json

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_TITLE_FONT = Font(bold=True, size=16)
_SECTION_FILL = PatternFill("solid", fgColor="191F28")
_SECTION_FONT = Font(color="FFFFFF", bold=True, size=11)
_LABEL_FILL = PatternFill("solid", fgColor="F2F4F6")
_LABEL_FONT = Font(bold=True, size=10)
_HINT_FONT = Font(color="8B95A1", size=9, italic=True)
_THIN = Side(style="thin", color="D5DAE0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# 종합 비교표에서 나란히 놓을 조사표 수(#1..#N). 편집 때 열 복사로 늘릴 수 있음.
_N_COMPARE = 3

_SYSTEM = (
    "당신은 한국 공공기관 현장조사 보고서 양식을 설계하는 전문가입니다. "
    "추출 항목 목록(이름·샘플값·제목 위계)을 보고, 보고서에 쓰기 좋은 구조를 제안하세요.\n"
    "규칙:\n"
    "- sections: 항목들을 의미 단위(예: 일반현황/위치/제원/조사결과)로 묶고 순서를 정한다. "
    "모든 항목을 정확히 한 번씩 배치한다(제목 위계 항목 제외 가능).\n"
    "- 각 항목의 label 은 보고서에 어울리는 짧은 한글 표기(원래 이름을 다듬는 정도, 의미 변경 금지).\n"
    "- numeric_fields: 샘플값이 수치(길이·높이·개수 등)라 합계/평균이 의미 있는 항목만.\n"
    "- report_title: 보고서 제목 한 줄(예: '하천 인공구조물 현장조사 보고서').\n"
    "- field 값은 준 항목 이름을 '그대로'(글자 하나 안 바꾸고) 사용한다."
)

_SCHEMA = {
    "type": "object",
    "properties": {
        "report_title": {"type": "string"},
        "sections": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "title": {"type": "string"},
                    "fields": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "field": {"type": "string"},
                                "label": {"type": "string"},
                            },
                            "required": ["field", "label"],
                            "additionalProperties": False,
                        },
                    },
                },
                "required": ["title", "fields"],
                "additionalProperties": False,
            },
        },
        "numeric_fields": {"type": "array", "items": {"type": "string"}},
    },
    "required": ["report_title", "sections", "numeric_fields"],
    "additionalProperties": False,
}


def ai_design(fields: list[str], sample: dict, title_fields: list[str]) -> dict:
    """AI 에게 설계(JSON)를 받는다. 항목 이름·샘플값·제목 위계를 전달."""
    from app import config
    from core import vision_extract

    payload = {
        "항목들": [{"이름": f, "샘플값": str(sample.get(f, ""))[:40],
                   "위계": ("제목" if f in title_fields else "일반")} for f in fields],
    }
    resp = vision_extract._create_with_retry(
        vision_extract._client(),
        model=config.VISION_MODEL,
        max_tokens=6000,
        system=_SYSTEM,
        messages=[{"role": "user", "content":
                   "다음 추출 항목으로 보고서 양식 구조를 설계해 주세요.\n\n"
                   + json.dumps(payload, ensure_ascii=False)}],
        output_config={"format": {"type": "json_schema", "schema": _SCHEMA}},
    )
    text = next((b.text for b in resp.content if b.type == "text"), "")
    return json.loads(text)


def _fallback_design(fields: list[str], sample: dict, title_fields: list[str]) -> dict:
    """AI 응답이 항목을 빠뜨렸거나 실패했을 때의 결정적 폴백 — 한 섹션에 전부."""
    body = [f for f in fields if f not in title_fields]
    numeric = []
    for f in body:
        s = str(sample.get(f, "")).replace(",", "").strip()
        try:
            float(s)
            numeric.append(f)
        except ValueError:
            pass
    return {"report_title": "현장조사 보고서",
            "sections": [{"title": "조사 항목",
                          "fields": [{"field": f, "label": f} for f in body]}],
            "numeric_fields": numeric}


def sanitize_design(design: dict, fields: list[str], sample: dict,
                    title_fields: list[str]) -> dict:
    """AI 설계를 검증·보정: 없는 항목 제거, 빠진 항목은 '기타' 섹션으로 추가."""
    body = [f for f in fields if f not in title_fields]
    known = set(body)
    out_sections, placed = [], set()
    for sec in (design.get("sections") or []):
        fs = []
        for it in (sec.get("fields") or []):
            f = (it.get("field") or "").strip()
            if f in known and f not in placed:
                fs.append({"field": f, "label": (it.get("label") or f).strip() or f})
                placed.add(f)
        if fs:
            out_sections.append({"title": (sec.get("title") or "항목").strip(), "fields": fs})
    missing = [f for f in body if f not in placed]
    if missing:
        out_sections.append({"title": "기타", "fields": [{"field": f, "label": f} for f in missing]})
    if not out_sections:
        return _fallback_design(fields, sample, title_fields)
    numeric = [f for f in (design.get("numeric_fields") or []) if f in known]
    if not numeric:   # AI가 안 줬으면 샘플값으로 자동 판별
        numeric = _fallback_design(fields, sample, title_fields)["numeric_fields"]
    return {"report_title": (design.get("report_title") or "현장조사 보고서").strip(),
            "sections": out_sections, "numeric_fields": numeric}


def build_draft_workbook(design: dict, title_fields: list[str], out_path: str) -> str:
    """설계(JSON) → 초안 xlsx 조립. 시트1 종합비교({항목#N}), 시트2 대상지별 1장."""
    wb = Workbook()

    # ---- 시트1: 종합 비교표 ----
    ws = wb.active
    ws.title = "종합비교"
    ws["A1"] = design["report_title"] + " — 종합 비교"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = ("※ AI 초안입니다. 자유롭게 편집한 뒤 프로그램에 다시 올리면 "
                "{항목명#번호} 자리에 각 조사표 값이 채워집니다. 열을 복사해 #4, #5…로 늘릴 수 있습니다.")
    ws["A2"].font = _HINT_FONT
    r = 4
    ws.cell(row=r, column=1, value="항목").fill = _SECTION_FILL
    ws.cell(row=r, column=1).font = _SECTION_FONT
    for n in range(1, _N_COMPARE + 1):
        c = ws.cell(row=r, column=1 + n)
        # 제목 위계 항목이 있으면 조사표 구분 헤더로 사용, 없으면 '조사표 N'
        c.value = ("{%s#%d}" % (title_fields[0], n)) if title_fields else f"조사표 {n}"
        c.fill = _SECTION_FILL
        c.font = _SECTION_FONT
        c.alignment = Alignment(horizontal="center")
    r += 1
    num_rows: dict[str, int] = {}
    for sec in design["sections"]:
        sc = ws.cell(row=r, column=1, value=f"▎{sec['title']}")
        sc.font = Font(bold=True, size=11)
        r += 1
        for it in sec["fields"]:
            ws.cell(row=r, column=1, value=it["label"]).fill = _LABEL_FILL
            ws.cell(row=r, column=1).font = _LABEL_FONT
            ws.cell(row=r, column=1).border = _BORDER
            for n in range(1, _N_COMPARE + 1):
                c = ws.cell(row=r, column=1 + n, value="{%s#%d}" % (it["field"], n))
                c.border = _BORDER
            if it["field"] in design["numeric_fields"]:
                num_rows[it["label"]] = r
            r += 1
    if num_rows:
        r += 1
        ws.cell(row=r, column=1, value="요약(자동 계산)").fill = _SECTION_FILL
        ws.cell(row=r, column=1).font = _SECTION_FONT
        ws.cell(row=r, column=2, value="평균").font = _LABEL_FONT
        ws.cell(row=r, column=3, value="합계").font = _LABEL_FONT
        r += 1
        last_col = get_column_letter(1 + _N_COMPARE)
        for label, row_i in num_rows.items():
            ws.cell(row=r, column=1, value=label).fill = _LABEL_FILL
            ws.cell(row=r, column=1).font = _LABEL_FONT
            ws.cell(row=r, column=2, value=f"=AVERAGE(B{row_i}:{last_col}{row_i})")
            ws.cell(row=r, column=3, value=f"=SUM(B{row_i}:{last_col}{row_i})")
            r += 1
    ws.column_dimensions["A"].width = 24
    for n in range(1, _N_COMPARE + 1):
        ws.column_dimensions[get_column_letter(1 + n)].width = 20

    # ---- 시트2: 대상지별 1장 보고서 ----
    ws2 = wb.create_sheet("대상지보고서")
    ws2["A1"] = design["report_title"]
    ws2["A1"].font = _TITLE_FONT
    if title_fields:
        ws2["A2"] = "{%s}" % title_fields[0]
        ws2["A2"].font = Font(bold=True, size=12, color="3182F6")
    ws2["A3"] = ("※ 이 시트는 조사표 1건당 1장씩 복사되어 채워집니다"
                 "(일괄 처리에서 '대상지별 시트 이름' 선택 시).")
    ws2["A3"].font = _HINT_FONT
    r = 5
    for sec in design["sections"]:
        c = ws2.cell(row=r, column=1, value=sec["title"])
        c.fill = _SECTION_FILL
        c.font = _SECTION_FONT
        ws2.cell(row=r, column=2).fill = _SECTION_FILL
        r += 1
        for it in sec["fields"]:
            ws2.cell(row=r, column=1, value=it["label"]).fill = _LABEL_FILL
            ws2.cell(row=r, column=1).font = _LABEL_FONT
            ws2.cell(row=r, column=1).border = _BORDER
            ws2.cell(row=r, column=2, value="{%s}" % it["field"]).border = _BORDER
            r += 1
        r += 1
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 34

    wb.save(out_path)
    return out_path


def make_draft(fields: list[str], sample: dict, title_fields: list[str],
               out_path: str) -> str:
    """AI 설계 → 검증·보정 → 초안 조립. AI 실패 시 결정적 폴백 설계로 진행."""
    try:
        design = ai_design(fields, sample, title_fields)
    except Exception:  # noqa: BLE001
        design = {}
    design = sanitize_design(design or {}, fields, sample, title_fields)
    return build_draft_workbook(design, title_fields, out_path)
