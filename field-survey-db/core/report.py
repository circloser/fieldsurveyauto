"""보고서 생성 — 사용자가 만든 엑셀 양식의 {필드명} 자리표시자를 추출값으로 채운다.

수식(=...)은 그대로 보존되어 엑셀에서 열면 자동 계산된다.
출력: 전체 요약(DB) 시트 + 파일마다 채워진 보고서 시트.
"""
from __future__ import annotations

import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_PH = re.compile(r"\{([^}]+)\}")

_HEADER_FILL = PatternFill("solid", fgColor="191F28")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)


def read_grid(path: str, max_rows: int = 120, max_cols: int = 40) -> dict:
    """양식 첫 시트를 격자(문자열)로 읽는다. 수식은 '=...', 자리표시자는 '{...}' 그대로."""
    wb = load_workbook(path)
    ws = wb.worksheets[0]
    nrows = min(ws.max_row or 1, max_rows)
    ncols = min(ws.max_column or 1, max_cols)
    cells = []
    for r in range(1, nrows + 1):
        row = []
        for cc in range(1, ncols + 1):
            v = ws.cell(row=r, column=cc).value
            row.append("" if v is None else str(v))
        cells.append(row)
    # 병합 정보(표시용)
    merges = [str(m) for m in ws.merged_cells.ranges]
    wb.close()
    return {"sheet": ws.title, "nrows": nrows, "ncols": ncols, "cells": cells, "merges": merges}


def save_with_edits(src_path: str, edits: dict, out_path: str) -> str:
    """원본 양식을 열어 편집된 셀만 덮어쓰고 저장(스타일·병합 보존)."""
    wb = load_workbook(src_path)
    ws = wb.worksheets[0]
    for key, val in (edits or {}).items():
        try:
            r, c = (int(x) for x in key.split(","))
        except ValueError:
            continue
        ws.cell(row=r, column=c).value = (val if val != "" else None)
    wb.save(out_path)
    return out_path


def list_placeholders(path: str) -> list[str]:
    """양식에서 쓰인 {필드명} 목록을 반환(사용자에게 어떤 필드가 필요한지 안내)."""
    wb = load_workbook(path)
    found: set[str] = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and not c.value.startswith("="):
                    for m in _PH.findall(c.value):
                        found.add(m.strip())
    wb.close()
    return sorted(found)


def _coerce(val):
    """숫자로 보이면 숫자로(수식 계산이 되도록), 아니면 문자 그대로."""
    s = str(val).strip()
    if s == "":
        return ""
    try:
        f = float(s.replace(",", ""))
        return int(f) if f.is_integer() else f
    except ValueError:
        return s


def _split_token(token: str) -> tuple[str, int | None]:
    """'항목#2' → ('항목', 2) ; '항목' → ('항목', None). N은 조사표 순번(1부터)."""
    if "#" in token:
        name, _, idx = token.rpartition("#")
        idx = idx.strip()
        if idx.isdigit():
            return name.strip(), int(idx)
    return token.strip(), None


def _value_for(token: str, records: list[dict], cur_idx: int):
    """{항목}=현재(cur_idx) 조사표, {항목#N}=N번째 조사표 값."""
    name, idx = _split_token(token)
    i = (idx - 1) if idx is not None else cur_idx
    return records[i].get(name, "") if 0 <= i < len(records) else ""


def _fill_ws(ws, records: list[dict], cur_idx: int) -> None:
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if not (isinstance(v, str) and not v.startswith("=") and "{" in v):
                continue
            m = _PH.fullmatch(v.strip())
            if m:  # 셀 전체가 {필드} 하나 → 숫자면 숫자로(수식 대상)
                c.value = _coerce(_value_for(m.group(1).strip(), records, cur_idx))
            else:   # 문장 속 자리표시자 → 문자 치환
                c.value = _PH.sub(lambda mm: str(_value_for(mm.group(1).strip(), records, cur_idx)), v)


def _template_has_indexed(ws) -> bool:
    """양식에 {항목#N} 형태(특정 순번 참조)가 있으면 True → 종합(단일) 보고서 모드."""
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and not c.value.startswith("="):
                for tok in _PH.findall(c.value):
                    if _split_token(tok)[1] is not None:
                        return True
    return False


def _sheet_name(base: str, existing: set[str]) -> str:
    base = re.sub(r'[\\/*?:\[\]]', "_", (base or "보고서"))[:28] or "보고서"
    name, i = base, 1
    while name in existing:
        i += 1
        name = f"{base[:24]}_{i}"
    return name


def record_sheet_base(rec: dict, sheet_name_field: str | None) -> str:
    """시트 이름의 원천 문자열 — 대상지명 같은 추출값 우선, 없으면 파일명."""
    if sheet_name_field:
        v = str(rec.get(sheet_name_field, "") or "").strip()
        if v:
            return v
    return str(rec.get("_파일명", "") or "")


def build_report_workbook(template_path: str, records: list[dict],
                          summary_fields: list[str], out_path: str,
                          max_reports: int = 300,
                          sheet_name_field: str | None = None) -> str:
    """양식을 파일마다 복사해 채우고, 앞에 요약(DB) 시트를 붙여 저장.

    sheet_name_field 가 있으면 그 추출값(예: 대상지명)으로 시트 이름을 짓는다.
    """
    wb = load_workbook(template_path)
    master = wb.worksheets[0]

    if _template_has_indexed(master):
        # 종합(단일) 보고서: 양식 한 장에 특정 순번을 참조해 채운다.
        #   {항목}=1번 조사표, {항목#N}=N번째 조사표(요약(DB)/파일 순서와 동일).
        _fill_ws(master, records, 0)
        master.title = _sheet_name("종합보고서", set())
    else:
        made = 0
        for idx, rec in enumerate(records):
            if made >= max_reports:
                break
            ws = wb.copy_worksheet(master)
            base = record_sheet_base(rec, sheet_name_field)
            ws.title = _sheet_name(base, set(wb.sheetnames) - {ws.title})
            _fill_ws(ws, records, idx)
            made += 1
        master.title = _sheet_name("＿양식원본", set(wb.sheetnames) - {master.title})

    # 요약(DB) 시트를 맨 앞에
    summ = wb.create_sheet("요약(DB)", 0)
    headers = ["파일명"] + summary_fields
    summ.append(headers)
    for col in range(1, len(headers) + 1):
        cell = summ.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    summ.freeze_panes = "A2"
    for rec in records:
        summ.append([rec.get("_파일명", "")] + [rec.get(f, "") for f in summary_fields])
    for i, h in enumerate(headers, start=1):
        summ.column_dimensions[get_column_letter(i)].width = max(10, min(28, len(str(h)) * 2 + 4))

    wb.save(out_path)
    return out_path
