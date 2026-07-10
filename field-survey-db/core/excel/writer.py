"""엑셀 DB 생성 — 서식별 시트 + 요약 마스터 시트.

열 순서는 스키마(SCHEMAS)를 단일 기준으로 삼는다(매핑과 열 순서가 어긋나지 않도록).
검수필요(플래그) 셀은 노란색으로 강조하고 사유를 메모로 단다.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.extraction.form_detector import FORM_A, FORM_C, FORM_D, FORM_LABELS_KO
from core.extraction.schema import SCHEMAS
from core.extraction.schema.spec import ANCHOR_NOT_FOUND
from core.pipeline import ExtractionResult

_ID_COLS = ["파일명", "보명칭", "보코드"]

_SHEET_TITLE = {FORM_A: "인공구조물", FORM_C: "어도"}
_SHEET_ORDER = [FORM_A, FORM_C]

# 요약 마스터에 모을 핵심 필드(있으면 채움)
_SUMMARY_FIELDS = ["재질", "보길이", "어도유형", "평균경사도"]

_HEADER_FILL = PatternFill("solid", fgColor="191F28")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_FLAG_FILL = PatternFill("solid", fgColor="FFF3BF")   # 검수필요(노랑)
_CENTER = Alignment(horizontal="center", vertical="center")


def _schema_fields(form_type: str) -> list[str]:
    return [spec.name for spec in SCHEMAS.get(form_type, [])]


def _style_header(ws, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
    ws.freeze_panes = "A2"


def _autosize(ws, headers: list[str]) -> None:
    for i, h in enumerate(headers, start=1):
        width = max(10, min(28, len(str(h)) * 2 + 4))
        ws.column_dimensions[get_column_letter(i)].width = width


def _write_form_sheet(wb: Workbook, form_type: str, records) -> None:
    fields = _schema_fields(form_type)
    headers = _ID_COLS + fields
    ws = wb.create_sheet(title=_SHEET_TITLE.get(form_type, FORM_LABELS_KO.get(form_type, form_type)))
    ws.append(headers)
    _style_header(ws, len(headers))

    for rec in records:
        row = [rec.source_file, rec.structure_name, rec.structure_code]
        row += [rec.values.get(f, "") for f in fields]
        ws.append(row)
        r = ws.max_row
        # 플래그 셀 강조
        for f, status in rec.flags.items():
            if f in fields:
                c = ws.cell(row=r, column=len(_ID_COLS) + fields.index(f) + 1)
                c.fill = _FLAG_FILL
                reason = "정렬 어긋남 의심(라벨 못 찾음)" if status == ANCHOR_NOT_FOUND else "값 비어있음/미기재"
                c.comment = Comment(f"검수필요: {reason}", "자동추출")
    _autosize(ws, headers)


def _write_fish_sheet(wb: Workbook, records) -> None:
    """어류(D) — 종별 1행(long 형식)."""
    ws = wb.create_sheet(title="어류")
    headers = ["파일명", "하천명", "보명칭", "보코드", "조사기관", "No", "종명", "개체수"]
    ws.append(headers)
    _style_header(ws, len(headers))
    for rec in records:
        river = rec.values.get("하천명", "")
        org = rec.values.get("조사기관", "")
        if not rec.table_rows:
            ws.append([rec.source_file, river, rec.structure_name, rec.structure_code, org, "", "", ""])
            continue
        for i, sp in enumerate(rec.table_rows, start=1):
            ws.append([
                rec.source_file, river, rec.structure_name, rec.structure_code, org,
                i, sp.get("종명", ""), sp.get("개체수", ""),
            ])
    _autosize(ws, headers)


def _write_summary(wb: Workbook, records) -> None:
    ws = wb.create_sheet(title="요약(보별)")
    headers = ["보코드", "보명칭", "하천명"] + _SUMMARY_FIELDS + ["검수필요"]
    ws.append(headers)
    _style_header(ws, len(headers))

    # 보코드 기준으로 병합(코드 없으면 파일+보명칭)
    by_code: dict[str, dict] = {}
    for rec in records:
        key = rec.structure_code or f"{rec.source_file}:{rec.structure_name}"
        agg = by_code.setdefault(key, {
            "보코드": rec.structure_code,
            "보명칭": rec.structure_name,
            "하천명": rec.river_name or rec.values.get("하천명", ""),
            "검수필요": 0,
        })
        agg["검수필요"] += len(rec.flags)
        for f in _SUMMARY_FIELDS:
            if not agg.get(f) and rec.values.get(f):
                agg[f] = rec.values[f]

    for agg in by_code.values():
        ws.append([
            agg.get("보코드", ""), agg.get("보명칭", ""), agg.get("하천명", ""),
            *[agg.get(f, "") for f in _SUMMARY_FIELDS],
            agg.get("검수필요", 0) or "",
        ])
    _autosize(ws, headers)


def write_excel(result: ExtractionResult, out_path: str) -> str:
    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    _write_summary(wb, result.records)  # 요약을 맨 앞에

    by_form: dict[str, list] = {}
    for rec in result.records:
        by_form.setdefault(rec.form_type, []).append(rec)

    for form_type in _SHEET_ORDER:
        if by_form.get(form_type):
            _write_form_sheet(wb, form_type, by_form[form_type])

    if by_form.get(FORM_D):
        _write_fish_sheet(wb, by_form[FORM_D])

    wb.save(out_path)
    return out_path
