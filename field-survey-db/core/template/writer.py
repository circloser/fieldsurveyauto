"""템플릿 기반 엑셀 출력 — 박스 순서 = 열 순서, 파일 하나당 한 행.

sheet_name_field 를 주면 요약 시트 뒤에 대상지(행)마다 개별 시트(항목|값)를 만든다.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.report import _sheet_name, record_sheet_base

_HEADER_FILL = PatternFill("solid", fgColor="191F28")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_CENTER = Alignment(horizontal="center", vertical="center")

# 이상치 칸 — 옅은 주황 배경 + 진한 주황 굵은 글씨 + 셀 메모(사유)
_OUT_FILL = PatternFill("solid", fgColor="FFE9D8")
_OUT_FONT = Font(color="C2410C", bold=True)


def _mark_outliers(ws, row_idx: int, row: dict, fields: list[str],
                   col_offset: int = 2) -> None:
    """행 dict의 '_이상치'({필드: 사유})를 읽어 해당 셀에 서식·메모를 입힌다."""
    flags = row.get("_이상치") or {}
    if not flags:
        return
    for j, f in enumerate(fields):
        reason = flags.get(f)
        if not reason:
            continue
        c = ws.cell(row=row_idx, column=col_offset + j)
        c.fill = _OUT_FILL
        c.font = _OUT_FONT
        c.comment = Comment(str(reason), "오토다타")


def write_bundle_excel(groups: list[dict], out_path: str) -> str:
    """서식별 시트로 나눠 저장(AI 번들 추출용).

    groups: [{"label": 시트이름, "fields": [열...], "rows": [{'_파일명':.., field:val}]}]
    """
    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거 후 서식별로 생성
    used: set[str] = set()
    for g in groups:
        title = _sheet_name(g.get("label") or "추출결과", used)
        used.add(title)
        ws = wb.create_sheet(title)
        headers = ["파일명"] + list(g["fields"])
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=1, column=col)
            c.fill = _HEADER_FILL
            c.font = _HEADER_FONT
            c.alignment = _CENTER
        ws.freeze_panes = "A2"
        for row in g["rows"]:
            ws.append([row.get("_파일명", "")] + [row.get(f, "") for f in g["fields"]])
            _mark_outliers(ws, ws.max_row, row, list(g["fields"]))
        for i, h in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(i)].width = max(10, min(30, len(str(h)) * 2 + 4))
    if not wb.sheetnames:  # 추출 결과가 하나도 없을 때
        wb.create_sheet("추출결과")
    wb.save(out_path)
    return out_path


def _style_header(ws, ncols: int) -> None:
    for col in range(1, ncols + 1):
        c = ws.cell(row=1, column=col)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = _CENTER
    ws.freeze_panes = "A2"


def write_template_excel(rows: list[dict], fields: list[str], out_path: str,
                         sheet_name_field: str | None = None,
                         group_field: str | None = None,
                         max_sheets: int = 300) -> str:
    """rows: [{'_파일명':..., field: value, ...}], fields: 박스 순서.

    group_field: 그 필드 값(예: 제목)이 같은 행끼리 묶어 값 이름의 시트로 분류.
    sheet_name_field: 행(대상지)마다 개별 시트(항목|값). 둘은 배타적으로 쓰인다.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "추출결과"
    headers = ["파일명"] + fields
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = _CENTER
    ws.freeze_panes = "A2"

    for row in rows:
        ws.append([row.get("_파일명", "")] + [row.get(f, "") for f in fields])
        _mark_outliers(ws, ws.max_row, row, fields)

    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(10, min(30, len(str(h)) * 2 + 4))

    # 제목(양식)별 분류 시트 — 같은 값의 행들을 묶어 그 값 이름의 시트로
    if group_field:
        groups: dict[str, list[dict]] = {}
        for row in rows:
            key = str(row.get(group_field, "") or "").strip() or "(제목없음)"
            groups.setdefault(key, []).append(row)
        for gname, grows in list(groups.items())[:max_sheets]:
            s = wb.create_sheet(_sheet_name(gname, set(wb.sheetnames)))
            s.append(headers)
            _style_header(s, len(headers))
            for row in grows:
                s.append([row.get("_파일명", "")] + [row.get(f, "") for f in fields])
                _mark_outliers(s, s.max_row, row, fields)
            for i, h in enumerate(headers, start=1):
                s.column_dimensions[get_column_letter(i)].width = max(10, min(30, len(str(h)) * 2 + 4))

    # 대상지별 개별 시트(항목|값 세로표) — 시트 이름 = 선택한 추출값
    if sheet_name_field:
        for n, row in enumerate(rows):
            if n >= max_sheets:
                break
            s = wb.create_sheet(_sheet_name(record_sheet_base(row, sheet_name_field),
                                            set(wb.sheetnames)))
            s.append(["항목", "값"])
            for col in (1, 2):
                c = s.cell(row=1, column=col)
                c.fill = _HEADER_FILL
                c.font = _HEADER_FONT
                c.alignment = _CENTER
            s.append(["파일명", row.get("_파일명", "")])
            flags = row.get("_이상치") or {}
            for f in fields:
                s.append([f, row.get(f, "")])
                if flags.get(f):
                    c = s.cell(row=s.max_row, column=2)
                    c.fill = _OUT_FILL
                    c.font = _OUT_FONT
                    c.comment = Comment(str(flags[f]), "오토다타")
            s.column_dimensions["A"].width = 22
            s.column_dimensions["B"].width = 34
            s.freeze_panes = "A2"

    wb.save(out_path)
    return out_path
