"""Phase 4 검증 — 엑셀 서식별 시트 + 요약 마스터 (AC-7, AC-8)."""
import pytest
from openpyxl import load_workbook

from core.excel.writer import write_excel
from core.pipeline import run


@pytest.fixture(scope="module")
def workbook(request, tmp_path_factory):
    fixture = request.path.parent / "fixtures" / "sample.hwpx"
    if not fixture.exists():
        pytest.skip("샘플 없음")
    res = run([str(fixture)])
    out = tmp_path_factory.mktemp("xlsx") / "out.xlsx"
    write_excel(res, str(out))
    return load_workbook(out)


def test_has_per_form_sheets(workbook):
    assert "인공구조물" in workbook.sheetnames
    assert "어도" in workbook.sheetnames


def test_has_summary_master_sheet(workbook):
    assert "요약(보별)" in workbook.sheetnames
    ws = workbook["요약(보별)"]
    assert ws.max_row >= 2  # 헤더 + 최소 1행


def test_form_a_row_populated(workbook):
    ws = workbook["인공구조물"]
    headers = [c.value for c in ws[1]]
    row = {h: c.value for h, c in zip(headers, ws[2])}
    assert row["보코드"] == "5220140009"
    assert row["재질"] == "콘크리트"
    assert row["보길이"] == "30"


def test_form_c_row_populated(workbook):
    ws = workbook["어도"]
    headers = [c.value for c in ws[1]]
    row = {h: c.value for h, c in zip(headers, ws[2])}
    assert row["어도유형"] == "아이스하버식"
    assert row["어도폭"] == "3.7"


def test_flagged_cell_highlighted(workbook):
    """검수필요 셀은 채움색(노랑)으로 강조되어야 한다(AC-10 엑셀측)."""
    ws = workbook["인공구조물"]
    headers = [c.value for c in ws[1]]
    col = headers.index("바닥보호공길이") + 1
    cell = ws.cell(row=2, column=col)
    assert cell.fill is not None and cell.fill.fgColor.rgb.endswith("FFF3BF")
