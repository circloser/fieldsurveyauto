"""PDF 픽셀박스 파이프라인 검증 — 자동제안 + 박스/앵커 추출."""
import pytest

from core.pdf_pipeline import apply_pixel_template, field_order, suggest_pixel_boxes
from core.pdf_reader import read_pdf


@pytest.fixture(scope="module")
def doc(request):
    fx = request.path.parent / "fixtures" / "sample.pdf"
    if not fx.exists():
        pytest.skip("PDF 샘플 없음")
    return read_pdf(str(fx))


def test_cell_detection(request):
    fx = request.path.parent / "fixtures" / "sample.pdf"
    if not fx.exists():
        pytest.skip("PDF 샘플 없음")
    from core.pdf_reader import detect_cells
    cells = detect_cells(str(fx), 0)
    assert len(cells) > 50  # 표 칸이 많이 잡혀야 한다
    assert any("하천명" in c.text for c in cells)


def test_cell_based_suggest_accurate(request, doc):
    fx = request.path.parent / "fixtures" / "sample.pdf"
    if not fx.exists():
        pytest.skip("PDF 샘플 없음")
    from core.pdf_pipeline import suggest_from_cells
    boxes = suggest_from_cells(str(fx), 0)
    assert boxes
    vals = apply_pixel_template(doc.pages, boxes)
    # 칸 기반은 핵심 필드가 정확해야 한다(단어 방식보다 우수)
    assert vals.get("하천명") == "해남천"
    assert vals.get("관리 기관") == "전남 해남군"
    assert vals.get("보 길이") == "30"
    assert vals.get("바닥보호공 길이") == "2.25"  # hwpx 방식이 실패했던 칸


def test_page_auto_matching(request):
    """9장 템플릿을 순서 섞인 3장 부분문서에 적용 → 내용으로 올바른 페이지 매칭."""
    import fitz

    fx = request.path.parent / "fixtures" / "sample.pdf"
    if not fx.exists():
        pytest.skip("PDF 샘플 없음")
    from core.pdf_pipeline import match_pages, suggest_cells_maximal
    from core.pdf_reader import read_pdf

    full = read_pdf(str(fx), ocr_scanned=False)
    tmpl = []
    for p in full.pages:
        tmpl += suggest_cells_maximal(str(fx), p.page_no)

    # 원본 3,2,4 페이지를 섞어 3장 부분문서 생성
    sub_path = fx.parent / "_test_subset.pdf"
    src = fitz.open(str(fx))
    sub = fitz.open()
    for pno in [3, 2, 4]:
        sub.insert_pdf(src, from_page=pno, to_page=pno)
    sub.save(str(sub_path))
    sub.close()
    src.close()
    try:
        sd = read_pdf(str(sub_path), ocr_scanned=False)
        pm = match_pages(tmpl, sd.pages)
        # 섞인 순서대로 올바르게 매칭되어야 한다
        assert pm.get(3) == 0
        assert pm.get(2) == 1
        assert pm.get(4) == 2
        # 3개 페이지만 매칭(입력에 3장뿐)
        assert len(pm) == 3
    finally:
        sub_path.unlink(missing_ok=True)


def _draw_form(path, rows, y_top):
    """간단 조사표 합성 — 라벨|값 2열 표를 y_top부터 그린다(칸 테두리 + 텍스트)."""
    import fitz

    doc = fitz.open()
    page = doc.new_page(width=400, height=500)
    font = fitz.Font("cjk")   # 한글 지원 내장 폰트
    tw = fitz.TextWriter(page.rect)
    x0, x1, x2, rh = 40, 160, 360, 28
    for i, (label, value) in enumerate(rows):
        y = y_top + i * rh
        page.draw_rect(fitz.Rect(x0, y, x1, y + rh), color=(0, 0, 0), width=0.8)
        page.draw_rect(fitz.Rect(x1, y, x2, y + rh), color=(0, 0, 0), width=0.8)
        tw.append((x0 + 6, y + 18), label, font=font, fontsize=10)
        if value:
            tw.append((x1 + 6, y + 18), value, font=font, fontsize=10)
    tw.write_text(page)
    doc.save(str(path))
    doc.close()


def test_organic_extraction_survives_row_insert(tmp_path):
    """조사표에 줄이 추가되고 위치가 밀려도 라벨 칸 기준(유기적)으로 값이 정확."""
    from core.pdf_pipeline import suggest_from_cells

    base = tmp_path / "form_base.pdf"
    _draw_form(base, [("하천명", "해남천"), ("관리기관", "전남 해남군"), ("보길이", "30")],
               y_top=60)
    boxes = suggest_from_cells(str(base), 0)   # 원본 위치로 만든 템플릿
    assert {b["field"] for b in boxes} >= {"하천명", "관리기관", "보길이"}

    # 변형본: 맨 위에 새 행 추가 + 전체 2행(56pt) 아래로 밀림 — 값도 다르게
    drift = tmp_path / "form_drift.pdf"
    _draw_form(drift, [("조사구분", "정기"), ("비고", "줄추가"),
                       ("하천명", "가곡천"), ("관리기관", "강원 삼척시"), ("보길이", "25")],
               y_top=116)
    dd = read_pdf(str(drift), ocr_scanned=False)

    organic = apply_pixel_template(dd.pages, boxes, pdf_path=str(drift))
    assert organic.get("하천명") == "가곡천"        # 밀려도 라벨 따라 정확
    assert organic.get("관리기관") == "강원 삼척시"
    assert organic.get("보길이") == "25"

    # 대조: 좌표만 쓰면(기존 방식) 밀린 만큼 어긋난다 → 유기적 방식의 필요성 입증
    naive = apply_pixel_template(dd.pages, boxes)
    assert naive.get("하천명") != "가곡천"


def test_detect_title(request):
    """상단 큰 글씨(제목)를 위계 다른 항목으로 감지."""
    fx = request.path.parent / "fixtures" / "sample.pdf"
    if not fx.exists():
        pytest.skip("PDF 샘플 없음")
    from core.pdf_pipeline import detect_title
    t0 = detect_title(str(fx), 0)
    assert t0 and "조사표" in t0["text"]        # 실제 서식 제목
    assert t0["y0"] < 200                        # 페이지 상단
    t2 = detect_title(str(fx), 2)
    assert t2 and t2["text"] != t0["text"]       # 서식이 다르면 제목도 다름


def test_group_by_title_sheets(tmp_path):
    """제목별 분류 — 같은 제목(양식) 행끼리 그 이름의 시트로 묶인다."""
    from openpyxl import load_workbook
    from core.template.writer import write_template_excel
    rows = [
        {"_파일명": "a.pdf", "제목": "인공구조물 조사표", "하천명": "해남천"},
        {"_파일명": "b.pdf", "제목": "어도 조사표", "하천명": "가곡천"},
        {"_파일명": "c.pdf", "제목": "인공구조물 조사표", "하천명": "삼척천"},
        {"_파일명": "d.pdf", "제목": "", "하천명": "무제천"},
    ]
    out = tmp_path / "grouped.xlsx"
    write_template_excel(rows, ["제목", "하천명"], str(out), group_field="제목")
    wb = load_workbook(out)
    assert wb.sheetnames[0] == "추출결과"                    # 전체 요약 유지
    assert "인공구조물 조사표" in wb.sheetnames
    assert "어도 조사표" in wb.sheetnames
    assert "(제목없음)" in wb.sheetnames                     # 빈 제목 폴백
    g = wb["인공구조물 조사표"]
    names = [g.cell(row=r, column=1).value for r in (2, 3)]
    assert names == ["a.pdf", "c.pdf"]                       # 같은 양식 2행 묶임


def test_multi_bundle_one_row_each(request):
    """한 파일에 같은 조사표가 2묶음(9쪽×2=18쪽) → 묶음마다 페이지 매핑 1개씩."""
    import fitz

    fx = request.path.parent / "fixtures" / "sample.pdf"
    if not fx.exists():
        pytest.skip("PDF 샘플 없음")
    from core.pdf_pipeline import match_bundles, suggest_cells_maximal

    full = read_pdf(str(fx), ocr_scanned=False)
    tmpl = []
    for p in full.pages:
        tmpl += suggest_cells_maximal(str(fx), p.page_no)

    dbl_path = fx.parent / "_test_double.pdf"
    src = fitz.open(str(fx))
    dbl = fitz.open()
    dbl.insert_pdf(src)
    dbl.insert_pdf(src)   # 같은 조사표 2묶음
    dbl.save(str(dbl_path))
    dbl.close(); src.close()
    try:
        dd = read_pdf(str(dbl_path), ocr_scanned=False)
        bundles = match_bundles(tmpl, dd.pages)
        assert len(bundles) == 2                       # 묶음 2개 = 행 2개
        n = len(full.pages)
        for tp, ip in bundles[0].items():
            assert ip == tp                            # 1묶음: 앞 9쪽
        for tp, ip in bundles[1].items():
            assert ip == tp + n                        # 2묶음: 뒤 9쪽
        # 단일 파일이면 묶음 1개(기존 동작 유지)
        assert len(match_bundles(tmpl, full.pages)) == 1
    finally:
        dbl_path.unlink(missing_ok=True)


def test_maximal_covers_all_cells(request, doc):
    fx = request.path.parent / "fixtures" / "sample.pdf"
    if not fx.exists():
        pytest.skip("PDF 샘플 없음")
    from core.pdf_pipeline import suggest_cells_maximal
    boxes = suggest_cells_maximal(str(fx), 0)
    # 최대 생성: 칸 대부분 커버(소수 제안보다 훨씬 많음)
    assert len(boxes) > 80
    # 위치 중복 없음
    keys = [(round(b["x0"]), round(b["y0"])) for b in boxes]
    assert len(keys) == len(set(keys))
    # 값 칸이 라벨 이름으로 잡힘 (중복 이름 있으므로 박스 단위로 확인)
    got = []
    for b in boxes:
        got.append(apply_pixel_template(doc.pages, [b]).get(b["field"], ""))
    assert "전남 해남군" in got and "해남천" in got


def test_suggest_is_conservative(doc):
    """보수적 제안: 명백한 데이터(숫자·좌표·체크) 값에만, 오검출 최소화."""
    boxes = suggest_pixel_boxes(doc.pages[0])
    # 과다 제안이 아니어야 한다(예전 100+ → 소수)
    assert len(boxes) <= 20
    vals = apply_pixel_template(doc.pages, boxes)
    # 제안된 값에는 숫자/기호가 포함(데이터성)
    joined = " ".join(v for v in vals.values() if v)
    assert any(ch.isdigit() for ch in joined)


def test_pixel_box_text_mode(doc):
    p = doc.pages[0]
    # 하천명 값 영역을 감싸는 좌표 박스(앵커 없이 좌표만)
    from core.pdf_reader import find_label_word
    lw = find_label_word(p, "해남천")
    box = {"order": 1, "field": "하천명", "page": 0, "mode": "text",
           "x0": lw.x0 - 3, "y0": lw.y0 - 3, "x1": lw.x1 + 3, "y1": lw.y1 + 3}
    assert apply_pixel_template(doc.pages, [box])["하천명"] == "해남천"


def test_bold_and_field_order(doc):
    boxes = [
        {"order": 2, "field": "기상", "page": 0, "mode": "bold",
         "anchor": {"label": "기상상태", "relation": "right"}, "use_anchor": True,
         "x0": 0, "y0": 0, "x1": 1, "y1": 1},
        {"order": 1, "field": "하천", "page": 0, "mode": "text",
         "anchor": {"label": "하천명", "relation": "right"}, "use_anchor": True,
         "x0": 0, "y0": 0, "x1": 1, "y1": 1},
    ]
    assert field_order(boxes) == ["하천", "기상"]
    vals = apply_pixel_template(doc.pages, boxes)
    assert vals["기상"] == "흐림"  # 굵게만
    assert vals["하천"] == "해남천"
