"""보고서 생성 검증 — {필드명} 채우기, 숫자 변환, 수식 보존, 요약+개별 시트."""
from openpyxl import Workbook, load_workbook

from core.report import (
    build_report_workbook,
    list_placeholders,
    read_grid,
    save_with_edits,
)


def _make_template(path):
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "조사 보고서"
    ws["A3"] = "하천명"; ws["B3"] = "{하천명}"
    ws["A4"] = "보길이"; ws["B4"] = "{보 길이}"
    ws["A5"] = "두배"; ws["B5"] = "=B4*2"
    ws["A6"] = "메모"; ws["B6"] = "조사: {조사일시}"
    wb.save(path)


def test_list_placeholders(tmp_path):
    tpl = tmp_path / "tpl.xlsx"
    _make_template(tpl)
    ph = list_placeholders(str(tpl))
    assert set(ph) == {"하천명", "보 길이", "조사일시"}


def test_read_grid_and_edit(tmp_path):
    tpl = tmp_path / "tpl.xlsx"
    _make_template(tpl)
    grid = read_grid(str(tpl))
    assert grid["nrows"] >= 6 and grid["ncols"] >= 2
    # B3 = row3,col2 = {하천명}
    assert grid["cells"][2][1] == "{하천명}"
    assert grid["cells"][4][1] == "=B4*2"  # 수식 보존
    # 편집: B3 값 바꾸고 새 셀 추가
    out = tmp_path / "edited.xlsx"
    save_with_edits(str(tpl), {"3,2": "{가곡천대체}", "7,1": "추가"}, str(out))
    g2 = read_grid(str(out))
    assert g2["cells"][2][1] == "{가곡천대체}"
    assert g2["cells"][6][0] == "추가"


def test_build_report_sheet_name_field(tmp_path):
    """시트 이름 = 추출값(대상지). 중복은 접미사, 금지문자 치환, 빈값은 파일명 폴백."""
    tpl = tmp_path / "tpl.xlsx"
    _make_template(tpl)
    recs = [
        {"_파일명": "a.pdf", "하천명": "해남천", "보 길이": "30", "조사일시": ""},
        {"_파일명": "b.pdf", "하천명": "해남천", "보 길이": "25", "조사일시": ""},   # 중복 이름
        {"_파일명": "c.pdf", "하천명": "가곡천[본류]/상류", "보 길이": "", "조사일시": ""},  # 금지문자
        {"_파일명": "d.pdf", "하천명": "", "보 길이": "", "조사일시": ""},           # 빈값 → 파일명
    ]
    out = tmp_path / "out.xlsx"
    build_report_workbook(str(tpl), recs, ["하천명"], str(out), sheet_name_field="하천명")
    wb = load_workbook(out)
    assert "해남천" in wb.sheetnames
    assert any(s.startswith("해남천_") for s in wb.sheetnames)      # 중복 → 접미사
    assert "가곡천[본류]/상류" not in wb.sheetnames                  # 금지문자 원문 없음
    assert any("가곡천" in s for s in wb.sheetnames)
    assert "d.pdf" in wb.sheetnames                                  # 빈값 → 파일명 폴백


def test_write_template_excel_per_site_sheets(tmp_path):
    """보고서 양식 없이도 대상지별 시트(항목|값)가 생긴다."""
    from core.template.writer import write_template_excel
    rows = [
        {"_파일명": "a.pdf", "하천명": "해남천", "보 길이": "30"},
        {"_파일명": "b.pdf", "하천명": "가곡천", "보 길이": "25"},
    ]
    out = tmp_path / "flat.xlsx"
    write_template_excel(rows, ["하천명", "보 길이"], str(out), sheet_name_field="하천명")
    wb = load_workbook(out)
    assert wb.sheetnames[0] == "추출결과"
    assert "해남천" in wb.sheetnames and "가곡천" in wb.sheetnames
    s = wb["해남천"]
    assert [s["A1"].value, s["B1"].value] == ["항목", "값"]
    assert [s["A2"].value, s["B2"].value] == ["파일명", "a.pdf"]
    assert [s["A3"].value, s["B3"].value] == ["하천명", "해남천"]
    # 옵션 미사용이면 기존과 동일(시트 1개)
    out2 = tmp_path / "flat2.xlsx"
    write_template_excel(rows, ["하천명"], str(out2))
    assert load_workbook(out2).sheetnames == ["추출결과"]


def test_build_report(tmp_path):
    tpl = tmp_path / "tpl.xlsx"
    _make_template(tpl)
    recs = [
        {"_파일명": "a.pdf", "하천명": "해남천", "보 길이": "30", "조사일시": "2026-06-17"},
        {"_파일명": "b.pdf", "하천명": "가곡천", "보 길이": "25", "조사일시": "2026-06-29"},
    ]
    out = tmp_path / "out.xlsx"
    build_report_workbook(str(tpl), recs, ["하천명", "보 길이"], str(out))
    wb = load_workbook(out)
    # 요약 + 파일별 보고서 시트
    assert "요약(DB)" in wb.sheetnames
    assert "a.pdf" in wb.sheetnames and "b.pdf" in wb.sheetnames
    # 값 채움 + 숫자 변환 + 수식 보존
    a = wb["a.pdf"]
    assert a["B3"].value == "해남천"
    assert a["B4"].value == 30 and isinstance(a["B4"].value, int)
    assert a["B5"].value == "=B4*2"
    assert a["B6"].value == "조사: 2026-06-17"
    # 요약표
    s = wb["요약(DB)"]
    assert [c.value for c in s[1]] == ["파일명", "하천명", "보 길이"]


def test_build_report_indexed_combined(tmp_path):
    """{항목#N} 이 있으면 종합(단일) 보고서 한 장 — 특정 순번 참조."""
    tpl = tmp_path / "combined.xlsx"
    wb = Workbook(); ws = wb.active
    ws["A1"] = "종합 비교표"
    ws["A2"] = "1번 하천"; ws["B2"] = "{하천명#1}"
    ws["A3"] = "2번 하천"; ws["B3"] = "{하천명#2}"
    ws["A4"] = "2번 보길이"; ws["B4"] = "{보길이#2}"
    ws["A5"] = "기본(1번)"; ws["B5"] = "{하천명}"
    wb.save(tpl)
    recs = [
        {"_파일명": "a.pdf", "하천명": "탄천", "보길이": "20"},
        {"_파일명": "b.pdf", "하천명": "해남천", "보길이": "30"},
    ]
    out = tmp_path / "combined_out.xlsx"
    build_report_workbook(str(tpl), recs, ["하천명"], str(out))
    wb2 = load_workbook(out)
    assert "종합보고서" in wb2.sheetnames        # 개별 시트 모드 아님
    assert "a.pdf" not in wb2.sheetnames and "b.pdf" not in wb2.sheetnames
    s = wb2["종합보고서"]
    assert s["B2"].value == "탄천"               # #1
    assert s["B3"].value == "해남천"             # #2
    assert s["B4"].value == 30                    # #2 보길이(숫자 변환)
    assert s["B5"].value == "탄천"               # 인덱스 없으면 1번
