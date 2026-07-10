"""Phase 7 검증 — 어류(D) 종명/개체수 목록 추출 + 엑셀 어류 시트."""
import pytest
from openpyxl import load_workbook

from core.excel.writer import write_excel
from core.extraction.form_detector import FORM_D
from core.extraction.mapper import map_block
from core.extraction.segmenter import segment
from core.parsers.hwpx_parser import parse_hwpx
from core.pipeline import run


@pytest.fixture(scope="module")
def fish_record(request):
    fixture = request.path.parent / "fixtures" / "sample.hwpx"
    if not fixture.exists():
        pytest.skip("샘플 없음")
    doc = parse_hwpx(str(fixture))
    for b in segment(doc):
        if b.form_type == FORM_D:
            return map_block(b, "sample.hwpx")
    pytest.skip("어류 서식 없음")


def test_species_extracted(fish_record):
    species = {s["종명"]: s["개체수"] for s in fish_record.table_rows}
    assert species["붕어"] == "2"
    assert species["피라미"] == "32"
    assert species["갈겨니"] == "22"
    # 목록 경계를 넘어 안내문을 긁지 않아야 한다.
    assert len(fish_record.table_rows) == 8
    assert all("■" not in s["종명"] for s in fish_record.table_rows)


def test_fish_sheet_in_excel(request, tmp_path):
    fixture = request.path.parent / "fixtures" / "sample.hwpx"
    if not fixture.exists():
        pytest.skip("샘플 없음")
    res = run([str(fixture)])
    out = tmp_path / "out.xlsx"
    write_excel(res, str(out))
    wb = load_workbook(out)
    assert "어류" in wb.sheetnames
    ws = wb["어류"]
    # 헤더 + 최소 8종 행
    names = [ws.cell(row=r, column=7).value for r in range(2, ws.max_row + 1)]
    assert "붕어" in names and "피라미" in names
