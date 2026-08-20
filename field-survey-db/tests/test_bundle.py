"""bundle 신뢰도 플래그 + 서식별 엑셀 — 네트워크 없이 순수 로직 검증."""
from openpyxl import load_workbook

from core.bundle import _flags_for, group_rows
from core.extraction.form_detector import FORM_A, FORM_UNKNOWN
from core.template.writer import write_bundle_excel
from core.vision_extract import items_to_dict


def test_flags_empty_and_format():
    vals = {"하천명": "탄천", "보길이": "20", "용도": "", "월류수심": "abc"}
    flags = _flags_for(FORM_A, vals)
    assert flags.get("용도") == "빈값"          # 빈 값
    assert flags.get("월류수심") == "형식오류"    # 숫자 필드인데 비숫자
    assert "하천명" not in flags                 # 정상 텍스트
    assert "보길이" not in flags                 # 정상 숫자


def test_items_to_dict_generic():
    # 범용 모드: 항목:값 배열 → dict, 빈 항목 제외, 중복 항목 유일화
    items = [
        {"항목": "하천명", "값": "탄천"},
        {"항목": "", "값": "무시"},          # 항목 없음 → 제외
        {"항목": "수온", "값": "18"},
        {"항목": "수온", "값": "19"},          # 중복 → '수온 (2)'
    ]
    d = items_to_dict(items)
    assert d["하천명"] == "탄천"
    assert d["수온"] == "18" and d["수온 (2)"] == "19"
    assert "무시" not in d.values() or "" not in d
    assert items_to_dict(None) == {}


def test_bundle_excel_sheets(tmp_path):
    out = tmp_path / "b.xlsx"
    groups = [
        {"label": "인공구조물", "fields": ["하천명", "보길이"],
         "rows": [{"_파일명": "a.pdf p1", "하천명": "탄천", "보길이": "20"}]},
        {"label": "어도", "fields": ["어도폭", "어도유형"],
         "rows": [{"_파일명": "a.pdf p3", "어도폭": "3.5", "어도유형": "계단식"}]},
    ]
    write_bundle_excel(groups, str(out))
    wb = load_workbook(out)
    assert wb.sheetnames == ["인공구조물", "어도"]
    ws = wb["인공구조물"]
    assert [c.value for c in ws[1]] == ["파일명", "하천명", "보길이"]
    assert [c.value for c in ws[2]] == ["a.pdf p1", "탄천", "20"]


def test_group_rows_known_and_generic_subclassify():
    # 정의된 서식 + 미상 서식을 '양식제목'으로 세분류(표기차는 병합)
    rows = [
        {"_파일명": "a p1", "form": FORM_A, "form_title": "", "label": "인공구조물",
         "values": {"하천명": "탄천"}},
        {"_파일명": "b p1", "form": FORM_UNKNOWN, "form_title": "수질 조사표",
         "label": "수질 조사표", "values": {"수온": "18"}},
        {"_파일명": "c p1", "form": FORM_UNKNOWN, "form_title": "수질  조사표",  # 공백차이
         "label": "수질  조사표", "values": {"수온": "19", "pH": "7"}},
        {"_파일명": "d p1", "form": FORM_UNKNOWN, "form_title": "토양 조사표",
         "label": "토양 조사표", "values": {"토성": "사질"}},
    ]
    groups = group_rows(rows)
    labels = [g["label"] for g in groups]
    assert "인공구조물" in labels
    assert labels.count("수질 조사표") == 1          # 표기 차이 병합 → 한 시트
    assert "토양 조사표" in labels
    water = next(g for g in groups if g["label"] == "수질 조사표")
    assert len(water["rows"]) == 2                    # 두 건 누적
    assert water["fields"] == ["수온", "pH"]          # 항목 합집합(등장 순서)
