"""4번 일괄 처리 기본 흐름 — 자동 대조 + 제목별 시트(제목 없으면 시트 하나)."""
import io
import json

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app
from tests.test_pdf_pipeline import _draw_form

client = TestClient(app)


def _boxes_for(pdf_path):
    from core.pdf_pipeline import suggest_from_cells
    return suggest_from_cells(str(pdf_path), 0)


def _apply(named_files, boxes):
    files = [("files", (name, open(path, "rb"), "application/pdf"))
             for name, path in named_files]
    try:
        return client.post(
            "/api/pdf/apply",
            data={"boxes": json.dumps(boxes),
                  "sheet_name_field": "__group_title__",
                  "auto_classify": "1"},
            files=files)
    finally:
        for _, (_, fh, _) in files:
            fh.close()


def test_default_apply_groups_by_title(tmp_path):
    """제목이 다른 조사표 → 제목 이름의 시트로 나뉜다(제목 박스 없어도 큰 글씨 감지)."""
    base = tmp_path / "base.pdf"
    _draw_form(base, [("하천명", "한천"), ("보길이", "30")], y_top=60, title="하천 조사표")
    a = tmp_path / "a.pdf"
    _draw_form(a, [("하천명", "가곡천"), ("보길이", "25")], y_top=60, title="하천 조사표")
    b = tmp_path / "b.pdf"
    _draw_form(b, [("하천명", "묵논습지"), ("보길이", "12")], y_top=60, title="습지 조사표")

    r = _apply([("a.pdf", a), ("b.pdf", b)], _boxes_for(base))
    assert r.status_code == 200
    d = r.json()
    assert d["auto_classify"] is True
    assert d["ok_count"] == 2
    assert d["forms"] == 2
    assert {g["form"] for g in d["by_form"]} == {"하천 조사표", "습지 조사표"}

    dl = client.get("/api/pdf/download")
    assert dl.status_code == 200
    wb = load_workbook(io.BytesIO(dl.content))
    assert set(wb.sheetnames) == {"하천 조사표", "습지 조사표"}


def test_default_apply_without_title_single_sheet(tmp_path):
    """제목이 아예 없으면 — 전부 시트 하나(추출결과)에 들어간다."""
    base = tmp_path / "base.pdf"
    _draw_form(base, [("하천명", "한천"), ("보길이", "30")], y_top=60)
    a = tmp_path / "a.pdf"
    _draw_form(a, [("하천명", "가곡천"), ("보길이", "25")], y_top=60)
    b = tmp_path / "b.pdf"
    _draw_form(b, [("하천명", "오십천"), ("보길이", "18")], y_top=60)

    r = _apply([("a.pdf", a), ("b.pdf", b)], _boxes_for(base))
    assert r.status_code == 200
    d = r.json()
    assert d["auto_classify"] is True
    assert d["ok_count"] == 2
    assert d["forms"] == 1
    assert d["by_form"][0]["form"] == "추출결과"

    dl = client.get("/api/pdf/download")
    wb = load_workbook(io.BytesIO(dl.content))
    assert wb.sheetnames == ["추출결과"]
    ws = wb["추출결과"]
    assert ws.max_row == 3  # 헤더 + 2행


class _FakeStore:
    def __init__(self, d):
        self.d = d

    def list_names(self):
        return list(self.d)

    def get(self, n):
        return self.d.get(n)


def test_classify_bundles_by_title_similarity(tmp_path, monkeypatch):
    """한 파일에 제목이 다른 조사표가 섞여 있으면 — 입력 제목이 템플릿 제목과
    유사한 양식으로 각각 분류되어 제목별 시트로 나뉜다(라벨이 같아도 제목이 가른다)."""
    import app.main as app_main

    # 템플릿 2개: 표 구조(라벨)는 같고 제목만 다름
    tpl_a = tmp_path / "tpl_하천.pdf"
    _draw_form(tpl_a, [("하천명", ""), ("보길이", "")], y_top=60, title="하천 조사표")
    tpl_b = tmp_path / "tpl_습지.pdf"
    _draw_form(tpl_b, [("하천명", ""), ("보길이", "")], y_top=60, title="습지 조사표")
    tpl_paths = {"하천": tpl_a, "습지": tpl_b}
    store = _FakeStore({
        "하천": {"name": "하천", "boxes": _boxes_for(tpl_a)},
        "습지": {"name": "습지", "boxes": _boxes_for(tpl_b)},
    })
    monkeypatch.setattr(app_main, "_TEMPLATES", store)
    monkeypatch.setattr(app_main, "_tpl_pdf_path", lambda name: tpl_paths[name])

    # 입력: 두 양식이 섞인 한 파일 (제목 끝 숫자까지 포함)
    import fitz
    p1 = tmp_path / "p1.pdf"
    _draw_form(p1, [("하천명", "가곡천"), ("보길이", "25")], y_top=60, title="하천 조사표 1")
    p2 = tmp_path / "p2.pdf"
    _draw_form(p2, [("하천명", "묵논습지"), ("보길이", "12")], y_top=60, title="습지 조사표 1")
    mixed = tmp_path / "mixed.pdf"
    m = fitz.open()
    for p in (p1, p2):
        src = fitz.open(str(p))
        m.insert_pdf(src)
        src.close()
    m.save(str(mixed))
    m.close()

    r = _apply([("mixed.pdf", mixed)], [])   # 현재 박스 없이 — 저장 템플릿만으로 분류
    assert r.status_code == 200
    d = r.json()
    assert d["ok_count"] == 2
    assert d["forms"] == 2
    # 시트 이름은 '템플릿 양식'의 제목 — 입력 제목("… 1")의 변형은 따라가지 않는다
    assert {g["form"] for g in d["by_form"]} == {"하천 조사표", "습지 조사표"}
    tpl_used = {m2["template"] for m2 in d["match_info"]}
    assert tpl_used == {"하천", "습지"}      # 제목 유사도로 각자 맞는 템플릿에 배정

    import io
    from openpyxl import load_workbook
    dl = client.get("/api/pdf/download")
    wb = load_workbook(io.BytesIO(dl.content))
    assert set(wb.sheetnames) == {"하천 조사표", "습지 조사표"}
    ws = wb["하천 조사표"]
    vals = [c.value for row in ws.iter_rows(min_row=2) for c in row]
    assert "가곡천" in vals                   # 값도 맞는 시트에 들어간다


def test_page_level_classify_title_only(tmp_path, monkeypatch):
    """페이지 단위 대조 — 표 라벨이 템플릿과 달라도(라벨 근거 0) 페이지 제목이
    템플릿 제목과 유사하면 그 양식으로 배정되어 제목별 시트로 정리된다."""
    import app.main as app_main

    tpl_a = tmp_path / "tpl_하천.pdf"
    _draw_form(tpl_a, [("하천명", ""), ("보길이", "")], y_top=60, title="하천 조사표")
    tpl_b = tmp_path / "tpl_습지.pdf"
    _draw_form(tpl_b, [("습지명", ""), ("수심", "")], y_top=60, title="습지 조사표")
    tpl_paths = {"하천": tpl_a, "습지": tpl_b}
    store = _FakeStore({
        "하천": {"name": "하천", "boxes": _boxes_for(tpl_a)},
        "습지": {"name": "습지", "boxes": _boxes_for(tpl_b)},
    })
    monkeypatch.setattr(app_main, "_TEMPLATES", store)
    monkeypatch.setattr(app_main, "_tpl_pdf_path", lambda name: tpl_paths[name])

    # 입력: 라벨이 다른(명칭/길이) 페이지들 — 제목만이 분류 근거
    import fitz
    p1 = tmp_path / "p1.pdf"
    _draw_form(p1, [("명칭", "가곡천"), ("길이", "25")], y_top=60, title="하천 조사표 5")
    p2 = tmp_path / "p2.pdf"
    _draw_form(p2, [("명칭", "묵논습지"), ("길이", "12")], y_top=60, title="습지 조사표 5")
    mixed = tmp_path / "mixed2.pdf"
    m = fitz.open()
    for p in (p1, p2):
        src = fitz.open(str(p))
        m.insert_pdf(src)
        src.close()
    m.save(str(mixed))
    m.close()

    r = _apply([("mixed2.pdf", mixed)], [])
    assert r.status_code == 200
    d = r.json()
    assert d["ok_count"] == 2
    assert d["forms"] == 2
    assert {g["form"] for g in d["by_form"]} == {"하천 조사표", "습지 조사표"}
    tpl_used = {m2["template"] for m2 in d["match_info"]}
    assert tpl_used == {"하천", "습지"}

    import io
    from openpyxl import load_workbook
    dl = client.get("/api/pdf/download")
    wb = load_workbook(io.BytesIO(dl.content))
    assert set(wb.sheetnames) == {"하천 조사표", "습지 조사표"}
    vals = [c.value for row in wb["하천 조사표"].iter_rows(min_row=2) for c in row]
    assert "가곡천" in vals   # 라벨이 달라도 좌표 폴백으로 값까지 추출


def test_booklet_template_splits_per_page(tmp_path, monkeypatch):
    """묶음집 템플릿(쪽마다 제목이 다른 소양식) — 입력 페이지 하나가 1행이 되고,
    쪽 제목별 시트로 나뉜다(사용자 실사례: 9쪽 묶음집 × 반복 세트)."""
    import app.main as app_main
    import fitz
    from core.pdf_pipeline import suggest_from_cells

    # 2쪽짜리 묶음집 템플릿: 쪽마다 제목·라벨이 다르다
    pa = tmp_path / "ta.pdf"
    _draw_form(pa, [("하천명", ""), ("보길이", "")], y_top=60, title="하천 조사표")
    pb = tmp_path / "tb.pdf"
    _draw_form(pb, [("습지명", ""), ("수심", "")], y_top=60, title="습지 조사표")
    tpl = tmp_path / "tpl_booklet.pdf"
    m = fitz.open()
    for p in (pa, pb):
        src = fitz.open(str(p)); m.insert_pdf(src); src.close()
    m.save(str(tpl)); m.close()
    boxes = suggest_from_cells(str(tpl), 0) + suggest_from_cells(str(tpl), 1)
    assert {int(b.get("page", 0)) for b in boxes} == {0, 1}

    store = _FakeStore({"묶음집": {"name": "묶음집", "boxes": boxes}})
    monkeypatch.setattr(app_main, "_TEMPLATES", store)
    monkeypatch.setattr(app_main, "_tpl_pdf_path", lambda name: tpl)

    # 입력: [하천, 습지] 세트 × 2 (4쪽) — 값은 세트마다 다르게
    pages = []
    for i, (river, wet) in enumerate([("가곡천", "묵논습지"), ("오십천", "산들습지")]):
        r = tmp_path / f"in_r{i}.pdf"
        _draw_form(r, [("하천명", river), ("보길이", "25")], y_top=60, title="하천 조사표 1")
        w = tmp_path / f"in_w{i}.pdf"
        _draw_form(w, [("습지명", wet), ("수심", "3")], y_top=60, title="습지 조사표 1")
        pages += [r, w]
    mixed = tmp_path / "mixed_booklet.pdf"
    m = fitz.open()
    for p in pages:
        src = fitz.open(str(p)); m.insert_pdf(src); src.close()
    m.save(str(mixed)); m.close()

    r = _apply([("mixed.pdf", mixed)], [])
    assert r.status_code == 200
    d = r.json()
    assert d["ok_count"] == 4                       # 페이지 1장 = 1행
    assert d["forms"] == 2
    # 시트 이름 = 템플릿 '쪽'의 제목 (입력 제목의 "… 1" 변형은 무시)
    assert {g["form"] for g in d["by_form"]} == {"하천 조사표", "습지 조사표"}
    assert all(g["count"] == 2 for g in d["by_form"])

    import io
    from openpyxl import load_workbook
    dl = client.get("/api/pdf/download")
    wb = load_workbook(io.BytesIO(dl.content))
    vals_r = [c.value for row in wb["하천 조사표"].iter_rows(min_row=2) for c in row]
    vals_w = [c.value for row in wb["습지 조사표"].iter_rows(min_row=2) for c in row]
    assert "가곡천" in vals_r and "오십천" in vals_r
    assert "묵논습지" in vals_w and "산들습지" in vals_w


def test_report_generate_from_batch_result(tmp_path):
    """5번 새 로직 — 4번 일괄 처리 결과를 보고서 양식에 채워 정리(재추출 없음)."""
    import io
    from openpyxl import Workbook, load_workbook

    # ① 4번 일괄 처리 실행(제목 있는 조사표 2건)
    base = tmp_path / "base.pdf"
    _draw_form(base, [("하천명", "한천"), ("보길이", "30")], y_top=60, title="하천 조사표")
    a = tmp_path / "a.pdf"
    _draw_form(a, [("하천명", "가곡천"), ("보길이", "25")], y_top=60, title="하천 조사표")
    b = tmp_path / "b.pdf"
    _draw_form(b, [("하천명", "오십천"), ("보길이", "18")], y_top=60, title="하천 조사표")
    r = _apply([("a.pdf", a), ("b.pdf", b)], _boxes_for(base))
    assert r.status_code == 200 and r.json()["ok_count"] == 2

    # ② 보고서 양식 업로드({하천명} 자리표시자)
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "하천 이름"
    ws["B1"] = "{하천명}"
    tpl = tmp_path / "rpt_tpl.xlsx"
    wb.save(str(tpl))
    with tpl.open("rb") as f:
        lr = client.post("/api/report/load", files={"file": ("rpt.xlsx", f,
                         "application/octet-stream")})
    rid = lr.json()["report_id"]

    # ③ 정리 실행 → 다운로드 (4번 재실행 없이)
    g = client.post("/api/report/generate", json={"report_id": rid, "edits": {}})
    assert g.status_code == 200, g.text
    assert g.json()["rows"] == 2
    dl = client.get("/api/report/result")
    assert dl.status_code == 200
    out = load_workbook(io.BytesIO(dl.content))
    assert out.sheetnames[0] == "요약(DB)"
    # 조사표마다 개별 시트 — 제목 값으로 시트 이름
    assert any("하천 조사표" in sn for sn in out.sheetnames[1:])
    filled = [ws2["B1"].value for sn in out.sheetnames[1:]
              for ws2 in [out[sn]] if ws2["B1"].value not in (None, "{하천명}")]
    assert "가곡천" in filled and "오십천" in filled


def test_outlier_cells_styled_in_excel(tmp_path):
    """4번 일괄 처리 엑셀 — 이상치 칸은 주황 서식 + 사유 메모로 표시된다."""
    import io
    from openpyxl import load_workbook

    base = tmp_path / "base.pdf"
    _draw_form(base, [("하천명", "한천"), ("보길이", "30")], y_top=60, title="하천 조사표")
    files = []
    for i, bo in enumerate(["25", "25", "25", "25", "349"]):   # 마지막이 이상치
        p = tmp_path / f"in{i}.pdf"
        _draw_form(p, [("하천명", f"천{i}"), ("보길이", bo)], y_top=60, title="하천 조사표")
        files.append((f"in{i}.pdf", p))

    r = _apply(files, _boxes_for(base))
    assert r.status_code == 200
    d = r.json()
    assert d["ok_count"] == 5
    assert d["outlier_count"] >= 1                    # 349가 잡힌다

    dl = client.get("/api/pdf/download")
    wb = load_workbook(io.BytesIO(dl.content))
    ws = wb[wb.sheetnames[0]]
    hdr = [c.value for c in ws[1]]
    col = hdr.index("보길이") + 1
    styled = []
    for row in ws.iter_rows(min_row=2):
        c = row[col - 1]
        if c.value == "349":
            styled.append((c.fill.fgColor.rgb, c.comment is not None))
    assert styled, "349 값 셀이 있어야 한다"
    rgb, has_comment = styled[0]
    assert str(rgb).endswith("FFE9D8")                # 주황 배경
    assert has_comment                                 # 사유 메모
    # 정상 값 셀은 서식 없음
    normal = next(row[col - 1] for row in ws.iter_rows(min_row=2)
                  if row[col - 1].value == "25")
    assert str(normal.fill.fgColor.rgb).lower() in ("00000000", "none")


def test_sheet_named_by_template_title_accumulates(tmp_path, monkeypatch):
    """시트 이름은 템플릿 양식의 제목 — 입력 제목이 차수 등으로 조금씩 달라도
    ('1차 하천 조사표', '2차 하천 조사표') 같은 양식이면 한 시트에 축적된다."""
    import io
    import fitz
    import app.main as app_main
    from openpyxl import load_workbook

    tpl = tmp_path / "tpl.pdf"
    _draw_form(tpl, [("하천명", ""), ("보길이", "")], y_top=60, title="하천 조사표")
    store = _FakeStore({"하천": {"name": "하천", "boxes": _boxes_for(tpl)}})
    monkeypatch.setattr(app_main, "_TEMPLATES", store)
    monkeypatch.setattr(app_main, "_tpl_pdf_path", lambda name: tpl)

    p1 = tmp_path / "p1.pdf"
    _draw_form(p1, [("하천명", "가곡천"), ("보길이", "25")], y_top=60, title="1차 하천 조사표")
    p2 = tmp_path / "p2.pdf"
    _draw_form(p2, [("하천명", "오십천"), ("보길이", "18")], y_top=60, title="2차 하천 조사표")
    mixed = tmp_path / "mixed_r.pdf"
    m = fitz.open()
    for p in (p1, p2):
        src = fitz.open(str(p)); m.insert_pdf(src); src.close()
    m.save(str(mixed)); m.close()

    r = _apply([("mixed.pdf", mixed)], [])
    assert r.status_code == 200
    d = r.json()
    assert d["ok_count"] == 2
    assert d["forms"] == 1                                  # 시트 하나로 축적
    assert d["by_form"][0]["form"] == "하천 조사표"          # 템플릿 제목이 시트명

    dl = client.get("/api/pdf/download")
    wb = load_workbook(io.BytesIO(dl.content))
    assert wb.sheetnames == ["하천 조사표"]
    ws = wb["하천 조사표"]
    hdr = [c.value for c in ws[1]]
    assert "하천명" in hdr and "보길이" in hdr               # 항목명 = 템플릿 기준
    vals = [c.value for row in ws.iter_rows(min_row=2) for c in row]
    assert "가곡천" in vals and "오십천" in vals             # 두 차수가 한 시트에


def test_unmatched_form_discarded(tmp_path, monkeypatch):
    """템플릿에 없는 양식은 버림 — 이름이 비슷한 형제 양식('…현장사진')이
    '…현장조사표' 시트에 억지로 섞이지 않는다(사용자 실사례 재현)."""
    import io
    import fitz
    import app.main as app_main
    from openpyxl import load_workbook

    # 템플릿: '인공구조물 현장조사표' 하나뿐 (사진 양식 없음)
    tpl = tmp_path / "tpl.pdf"
    _draw_form(tpl, [("하천명", ""), ("보길이", "")], y_top=60, title="인공구조물 현장조사표")
    store = _FakeStore({"조사표": {"name": "조사표", "boxes": _boxes_for(tpl)}})
    monkeypatch.setattr(app_main, "_TEMPLATES", store)
    monkeypatch.setattr(app_main, "_tpl_pdf_path", lambda name: tpl)

    # 입력: 조사표 1쪽 + 사진 1쪽(제목이 비슷하지만 다른 양식, 라벨도 다름)
    p1 = tmp_path / "p1.pdf"
    _draw_form(p1, [("하천명", "가곡천"), ("보길이", "25")], y_top=60,
               title="인공구조물 현장조사표 1")
    p2 = tmp_path / "p2.pdf"
    _draw_form(p2, [("보 전경", "사진1"), ("상류 방향", "사진2")], y_top=60,
               title="인공구조물 현장사진")
    mixed = tmp_path / "mixed_d.pdf"
    m = fitz.open()
    for p in (p1, p2):
        src = fitz.open(str(p)); m.insert_pdf(src); src.close()
    m.save(str(mixed)); m.close()

    r = _apply([("mixed.pdf", mixed)], [])
    assert r.status_code == 200
    d = r.json()
    assert d["ok_count"] == 1                       # 조사표만 추출
    assert d["forms"] == 1
    assert d["by_form"][0]["form"] == "인공구조물 현장조사표"
    # 버림 보고: 사진 양식의 제목이 확인 결과로 표시된다
    assert d["discarded"] == [{"title": "인공구조물 현장사진", "pages": 1}]

    dl = client.get("/api/pdf/download")
    wb = load_workbook(io.BytesIO(dl.content))
    assert wb.sheetnames == ["인공구조물 현장조사표"]
    vals = [c.value for row in wb["인공구조물 현장조사표"].iter_rows(min_row=2)
            for c in row]
    assert "가곡천" in vals
    assert "사진1" not in vals and "인공구조물 현장사진" not in vals  # 혼입 없음


def test_wrong_number_sibling_discarded(tmp_path, monkeypatch):
    """꼬리 번호가 다른 양식('조사표 1' 입력 vs '조사표 2' 템플릿만 존재)은
    표 구조(라벨)가 같아도 다른 양식으로 확인하고 버린다."""
    import fitz
    import app.main as app_main

    tpl = tmp_path / "tpl2.pdf"
    _draw_form(tpl, [("하천명", ""), ("보길이", "")], y_top=60, title="하천 조사표 2")
    store = _FakeStore({"하천템플릿": {"name": "하천템플릿", "boxes": _boxes_for(tpl)}})
    monkeypatch.setattr(app_main, "_TEMPLATES", store)
    monkeypatch.setattr(app_main, "_tpl_pdf_path", lambda name: tpl)

    p1 = tmp_path / "in1.pdf"
    _draw_form(p1, [("하천명", "가곡천"), ("보길이", "25")], y_top=60, title="하천 조사표 1")

    r = _apply([("in1.pdf", p1)], [])
    d = r.json()
    if r.status_code == 200:
        assert d.get("ok_count", 0) == 0
        assert d["discarded"] == [{"title": "하천 조사표 1", "pages": 1}]
    else:  # 전부 버려져 처리 행이 0이면 400 + 버림 안내도 허용
        assert "버림" in (d.get("error") or "") or d.get("failed")
