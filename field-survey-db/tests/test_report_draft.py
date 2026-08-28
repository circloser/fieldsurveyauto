"""AI 보고서 초안 — 설계 보정(sanitize)과 xlsx 조립(build)의 결정적 부분 검증."""
from openpyxl import load_workbook

from core.report_draft import build_draft_workbook, sanitize_design


FIELDS = ["제목", "하천명", "보 길이", "보 높이", "비고"]
SAMPLE = {"제목": "인공구조물 조사표", "하천명": "해남천", "보 길이": "30",
          "보 높이": "1.5", "비고": "특이사항 없음"}


def test_sanitize_fixes_ai_output():
    """없는 항목 제거·빠진 항목 '기타' 추가·제목 위계 제외."""
    ai = {"report_title": "하천 조사 보고서",
          "sections": [{"title": "제원", "fields": [
              {"field": "보 길이", "label": "보 길이(m)"},
              {"field": "없는항목", "label": "유령"},        # 제거돼야 함
              {"field": "보 길이", "label": "중복"},          # 중복 배치 제거
          ]}],
          "numeric_fields": ["보 길이", "없는항목"]}
    d = sanitize_design(ai, FIELDS, SAMPLE, ["제목"])
    placed = [f["field"] for s in d["sections"] for f in s["fields"]]
    assert placed.count("보 길이") == 1
    assert "없는항목" not in placed
    assert set(placed) == {"하천명", "보 길이", "보 높이", "비고"}   # 제목 제외 전부 정확히 1회
    assert d["sections"][-1]["title"] == "기타"                      # 빠진 항목 수용
    assert d["numeric_fields"] == ["보 길이"]


def test_sanitize_falls_back_on_empty():
    d = sanitize_design({}, FIELDS, SAMPLE, ["제목"])
    placed = [f["field"] for s in d["sections"] for f in s["fields"]]
    assert set(placed) == {"하천명", "보 길이", "보 높이", "비고"}
    assert "보 길이" in d["numeric_fields"]      # 샘플값이 숫자면 수식 대상 자동 판별
    assert "하천명" not in d["numeric_fields"]


def test_build_draft_workbook(tmp_path):
    d = sanitize_design({}, FIELDS, SAMPLE, ["제목"])
    out = tmp_path / "draft.xlsx"
    build_draft_workbook(d, ["제목"], str(out))
    wb = load_workbook(out)
    assert wb.sheetnames == ["종합비교", "대상지보고서"]
    all1 = "|".join(str(c.value) for row in wb["종합비교"].iter_rows() for c in row if c.value)
    assert "{하천명#1}" in all1 and "{하천명#3}" in all1     # 비교 열 #1..#3
    assert "{제목#1}" in all1                                 # 제목 위계 = 조사표 구분 헤더
    assert "=AVERAGE(" in all1 and "=SUM(" in all1            # 수식 요약 블록
    all2 = "|".join(str(c.value) for row in wb["대상지보고서"].iter_rows() for c in row if c.value)
    assert "{하천명}" in all2 and "{보 길이}" in all2         # 1장 양식 자리표시자
    assert "{제목}" in all2
    # 초안이 기존 채움 엔진(read/list)과 호환되는지
    from core.report import list_placeholders
    ph = set(list_placeholders(str(out)))
    assert {"하천명", "보 길이", "하천명#1"} <= ph
